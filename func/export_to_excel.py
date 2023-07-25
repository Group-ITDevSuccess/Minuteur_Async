from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from datetime import date
from openpyxl.styles import Alignment
from datetime import datetime
import os


def get_age(date_string):
    """Calculates the age in days between a given date and today's date"""
    today = date.today()
    delta = today - date_string.date()  # Convertir en objet date
    return delta.days


def get_type(type_value):
    type_mapping = {
        0: "Client",
        1: "Fournisseur",
        3: "Autre"
    }
    return type_mapping.get(type_value, "")


def export_to_excel(df, objet):
    # Obtention de la date actuelle
    current_date = datetime.now().strftime("%Y-%m-%d")
    # Créer le répertoire si nécessaire
    directory = "data"
    os.makedirs(directory, exist_ok=True)

    workbook = Workbook()

    # Feuille pour les clients
    clients_sheet = workbook.active
    clients_sheet.title = "Clients"

    # Feuille pour les fournisseurs
    fournisseurs_sheet = workbook.create_sheet(title="Fournisseurs")

    # Feuille pour les autres
    autres_sheet = workbook.create_sheet(title="Autres")

    # Création d'un style de bordure pour les cellules
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # Création d'un style de remplissage pour l'en-tête
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Création d'un style de police gras pour l'en-tête
    header_font = Font(bold=True)

    # En-têtes des colonnes
    headers = ["TYPE", "CODE", "INTITULE", "NON ECHU", "30", "60", "90", "90+", "SOLDES"]

    # En-têtes pour les clients
    for col_index, header in enumerate(headers, start=1):
        cell = clients_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes pour les fournisseurs
    for col_index, header in enumerate(headers, start=1):
        cell = fournisseurs_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes pour les autres
    for col_index, header in enumerate(headers, start=1):
        cell = autres_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dictionnaires pour stocker les sommes des valeurs
    clients = {}
    fournisseurs = {}
    autres = {}

    for index, row in df.iterrows():
        key = (get_type(row["TYPE"]), row["CODE"], row["INTITULE"])
        echeance = row["ECHEANCE"]
        solde = row["SOLDE"]

        if key[0] == "Client":
            if key not in clients:
                clients[key] = {
                    "NON_ECHU": 0,
                    "30": 0,
                    "60": 0,
                    "90": 0,
                    "90+": 0,
                    "SOLDES": 0
                }
            age = get_age(echeance)
            if age <= 0 or echeance.date() == date(1753, 1, 1):
                clients[key]["NON_ECHU"] += solde
            elif 0 < age <= 30:
                clients[key]["30"] += solde
            elif 30 < age <= 60:
                clients[key]["60"] += solde
            elif 60 < age <= 90:
                clients[key]["90"] += solde
            else:
                clients[key]["90+"] += solde

            clients[key]["SOLDES"] += solde
        elif key[0] == "Fournisseur":
            if key not in fournisseurs:
                fournisseurs[key] = {
                    "NON_ECHU": 0,
                    "30": 0,
                    "60": 0,
                    "90": 0,
                    "90+": 0,
                    "SOLDES": 0
                }
            age = get_age(echeance)
            if age <= 0 or echeance.date() == date(1753, 1, 1):
                fournisseurs[key]["NON_ECHU"] += solde
            elif 0 < age <= 30:
                fournisseurs[key]["30"] += solde
            elif 30 < age <= 60:
                fournisseurs[key]["60"] += solde
            elif 60 < age <= 90:
                fournisseurs[key]["90"] += solde
            else:
                fournisseurs[key]["90+"] += solde

            fournisseurs[key]["SOLDES"] += solde
        else:
            if key not in autres:
                autres[key] = {
                    "NON_ECHU": 0,
                    "30": 0,
                    "60": 0,
                    "90": 0,
                    "90+": 0,
                    "SOLDES": 0
                }
            age = get_age(echeance)
            if age <= 0 or echeance.date() == date(1753, 1, 1):
                autres[key]["NON_ECHU"] += solde
            elif 0 < age <= 30:
                autres[key]["30"] += solde
            elif 30 < age <= 60:
                autres[key]["60"] += solde
            elif 60 < age <= 90:
                autres[key]["90"] += solde
            else:
                autres[key]["90+"] += solde

            autres[key]["SOLDES"] += solde

    row_index = 2

    # Remplissage des données pour les clients
    for key, values in clients.items():
        for col_index, value in enumerate(key, start=1):
            clients_sheet.cell(row=row_index, column=col_index, value=value)

        for col_index, value in enumerate(values.values(), start=len(key) + 1):
            clients_sheet.cell(row=row_index, column=col_index, value=value)

        # Appliquer le style de bordure aux cellules de la ligne
        for column in range(1, clients_sheet.max_column + 1):
            clients_sheet.cell(row=row_index, column=column).border = border_style

        row_index += 1

    row_index = 2

    # Remplissage des données pour les fournisseurs
    for key, values in fournisseurs.items():
        for col_index, value in enumerate(key, start=1):
            fournisseurs_sheet.cell(row=row_index, column=col_index, value=value)

        for col_index, value in enumerate(values.values(), start=len(key) + 1):
            fournisseurs_sheet.cell(row=row_index, column=col_index, value=value)

        # Appliquer le style de bordure aux cellules de la ligne
        for column in range(1, fournisseurs_sheet.max_column + 1):
            fournisseurs_sheet.cell(row=row_index, column=column).border = border_style

        row_index += 1

    row_index = 2

    # Remplissage des données pour les autres
    for key, values in autres.items():
        for col_index, value in enumerate(key, start=1):
            autres_sheet.cell(row=row_index, column=col_index, value=value)

        for col_index, value in enumerate(values.values(), start=len(key) + 1):
            autres_sheet.cell(row=row_index, column=col_index, value=value)

        # Appliquer le style de bordure aux cellules de la ligne
        for column in range(1, autres_sheet.max_column + 1):
            autres_sheet.cell(row=row_index, column=column).border = border_style

        row_index += 1

    # Ajuster la largeur des colonnes pour s'adapter au contenu pour toutes les feuilles
    for sheet in workbook:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Sauvegarder le fichier Excel
    filename = f"{objet}_{current_date}.xlsx"
    pathname = f"data/{filename}"
    workbook.save(pathname)
    return pathname
